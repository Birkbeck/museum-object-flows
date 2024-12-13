import openpyxl


class ExcelWriter:
    """This class generates an excel file containing tables.
    Use save_sheets to translate a dictionary of key: table pairs into an xlsx file.
    Each worksheet in the file will be titled with the key.
    """

    def __init__(self, file_name: str):
        self.file_name = file_name

    def save_sheets(self, replacements: dict):
        workbook = openpyxl.Workbook()
        for replacement, table in replacements.items():
            worksheet = workbook.create_sheet(title=replacement)
            for col_index, col_name in enumerate(table.columns.keys()):
                column = table.columns[col_name]
                worksheet.cell(row=1, column=col_index + 1, value=column.name)
            for row_index, row in enumerate(table):
                for col_index, col_name in enumerate(table.columns.keys()):
                    column = table.columns[col_name]
                    worksheet.cell(
                        row=row_index + 2,
                        column=col_index + 1,
                        value=column.to_string(row[column.name]),
                    )
        workbook.save(self.file_name)
